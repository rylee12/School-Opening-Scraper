import logging
import requests
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import pandas as pd
from datetime import date
from datetime import datetime
import re


def main():
    logging.basicConfig(filename='app.log', filemode='a', format='%(asctime)s - %(message)s', level=logging.INFO)
    # Get html of page
    url = "http://education.ohio.gov/Topics/Reset-and-Restart"
    html = requests.get(url).content
    soup = BeautifulSoup(html, 'html.parser')
    modifiedDate = download_xslx(soup)
    logging.info("Received Ohio Data", exc_info=False);
    copy_to_new_csv(soup, modifiedDate)
    logging.info("Wrote Ohio Data", exc_info=False);


def download_xslx(soup):
    # Get the most recent update link
    path = soup.select_one('div[id="main-content"]').find("a", string="this data compilation")['href']
    date = path[-10:-6]
    dataUrl = "http://education.ohio.gov" + str(path)
    # Retrieve cvs file
    originalFile = requests.get(dataUrl)
    open('temp/OhioOriginal.xlsx', 'wb').write(originalFile.content)
    return date


def copy_to_new_csv(soup, modifiedDate):
    wb = load_workbook('temp/OhioOriginal.xlsx')
    districtSheet = wb['Model']
    inputRow = 0
    df = pd.DataFrame(
        columns=['district irn', 'district', 'county',
                 'current model', 'report date', 'date scraped'])

    for row in districtSheet.iter_rows(values_only=True):
        if inputRow == 0:  # Skip column headers
            inputRow += 1
            continue
        districtIRN = row[0]
        districtName = row[1]
        countyName = row[2]
        currentModel = row[3]
        if districtName is None:  # End of input file
            break

        newRow = pd.Series(data={'district irn': districtIRN, 'district': districtName,
                                         'county': countyName, 'current model': currentModel,
                                         'report date': modifiedDate, 'date scraped': date.today()})

        df = df.append(newRow, ignore_index=True)

        inputRow += 1  # End for
    
    # Get last updated data
    text = soup.get_text()
    last_updated = re.search(r'(\(Map\s*updated\s*)([A-Za-z0-9\,\s+]+)(\))', text).group(2)
    df["last_updated"] = last_updated
    df.to_csv('out/OH_' + datetime.now().strftime('%Y%m%d') + '.csv', index=False)  # Copy dataframe to CSV

#main()
